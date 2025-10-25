"""Excel file generation utilities."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def identify_rollup_rows(df):
    """Identify rollup rows based on naming patterns and Row_Type column.

    Args:
        df: pandas DataFrame

    Returns:
        list: List of row indices that are rollup/summary rows
    """
    rollup_indicators = ['total', 'gross', 'effective', 'net operating income', 'noi']
    rollup_rows = []

    # Check if Row_Type column exists
    if 'Row_Type' in df.columns:
        for idx in df.index:
            try:
                row_type = df.loc[idx, 'Row_Type']
                # Handle case where loc returns a Series
                if isinstance(row_type, pd.Series):
                    row_type = row_type.iloc[0] if len(row_type) > 0 else None
                if pd.notna(row_type) and str(row_type).strip().upper() == 'ROLLUP':
                    rollup_rows.append(idx)
            except (KeyError, IndexError):
                continue
    else:
        # Fallback: identify by naming patterns
        first_col = df.columns[0]
        for idx in df.index:
            try:
                value = df.loc[idx, first_col]
                # Handle case where loc returns a Series
                if isinstance(value, pd.Series):
                    value = value.iloc[0] if len(value) > 0 else None
                if pd.notna(value):
                    value_lower = str(value).lower().strip()
                    if any(indicator in value_lower for indicator in rollup_indicators):
                        rollup_rows.append(idx)
            except (KeyError, IndexError):
                continue

    return rollup_rows


def add_rollup_formulas(df, ws, rollup_rows):
    """Add Excel formulas for rollup rows that sum their component rows.

    Args:
        df: pandas DataFrame
        ws: openpyxl worksheet
        rollup_rows: List of row indices that are rollup rows
    """
    # Find label column names (usually first non-Row_Type column that contains text)
    label_cols = ['Row_Type', 'INCOME', 'Category', 'EXPENSES', 'Subcategory', 'Description', 'Item']

    # Find numeric columns (columns with currency/number patterns, exclude Row_Type, labels, and Notes)
    numeric_cols = []
    for col in df.columns:
        # Skip if it's a label column or Notes
        if col in label_cols or col in ['Notes', 'notes', 'Note']:
            continue
        # Check if column contains numeric-looking data (currency signs, numbers, percentages)
        sample_values = df[col].dropna().head(5).astype(str)
        if any(
            any(char in str(val) for char in ['$', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '%', ','])
            for val in sample_values
        ):
            numeric_cols.append(col)

    for rollup_idx in rollup_rows:
        # Find the range of detail rows
        # Strategy: Look forward first (for breakdowns that come AFTER the total row)
        # If none found, look backward (for totals that come AFTER detail rows)
        start_detail_idx = None
        end_detail_idx = None

        # Try looking forward first (for cases like "Total Other Income" followed by "Parking", "Utility", etc.)
        for i in range(rollup_idx + 1, len(df)):
            if 'Row_Type' in df.columns:
                row_type = df.at[i, 'Row_Type']
                if pd.notna(row_type):
                    row_type_str = str(row_type).strip().upper()
                    if row_type_str == 'DETAIL':
                        if start_detail_idx is None:
                            start_detail_idx = i
                        end_detail_idx = i
                    elif row_type_str in ['ROLLUP', 'HEADER']:
                        # Stop at next rollup or header
                        break

        # If no forward details found, look backwards (for cases like detail rows followed by "Total")
        if start_detail_idx is None:
            for i in range(rollup_idx - 1, -1, -1):
                if 'Row_Type' in df.columns:
                    row_type = df.at[i, 'Row_Type']
                    if pd.notna(row_type):
                        row_type_str = str(row_type).strip().upper()
                        if row_type_str == 'DETAIL':
                            if end_detail_idx is None:
                                end_detail_idx = i
                            start_detail_idx = i
                        elif row_type_str in ['ROLLUP', 'HEADER']:
                            # Stop at previous rollup or header
                            break

        if start_detail_idx is not None and end_detail_idx is not None:
            # Add formulas for each numeric column ONLY
            for col_idx, col_name in enumerate(df.columns):
                if col_name in numeric_cols:
                    # Excel row is +2 (1-indexed + header row)
                    excel_row = rollup_idx + 2
                    excel_start_row = start_detail_idx + 2
                    excel_end_row = end_detail_idx + 2

                    # Convert column index to Excel column letter
                    excel_col = get_column_letter(col_idx + 1)

                    # Create SUM formula
                    formula = f"=SUM({excel_col}{excel_start_row}:{excel_col}{excel_end_row})"
                    ws.cell(row=excel_row, column=col_idx + 1, value=formula)


def extract_general_notes(df):
    """Extract notes that are not tied to specific rows (for separate Notes tab).

    Args:
        df: pandas DataFrame

    Returns:
        list: List of general notes
    """
    general_notes = []

    if 'Notes' in df.columns:
        # Look for notes in header or standalone rows
        for idx in df.index:
            note = df.at[idx, 'Notes']
            if pd.notna(note):
                # Check if this is a general note (in a HEADER row or standalone)
                if 'Row_Type' in df.columns:
                    row_type = df.at[idx, 'Row_Type']
                    if pd.notna(row_type) and str(row_type).strip().upper() == 'HEADER':
                        general_notes.append(note)

    return general_notes


def save_excel_incremental(tables, output_path, current_page, total_pages):
    """Save progress to Excel file incrementally.

    Args:
        tables: List of table dictionaries
        output_path: Path to save Excel file
        current_page: Current page number being processed
        total_pages: Total number of pages
    """
    try:
        print(f"  💾 Saving progress... ({current_page}/{total_pages} pages processed)")
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for idx, table_data in enumerate(tables, start=1):
            df = table_data['dataframe']
            page_num = table_data['page']
            table_num = table_data['table']

            if len(tables) == 1:
                sheet_name = "Sheet1"
            else:
                sheet_name = f"Page{page_num}_Table{table_num}"

            if len(sheet_name) > 31:
                sheet_name = f"P{page_num}_T{table_num}"

            if sheet_name in wb.sheetnames:
                sheet_name = f"{sheet_name}_{idx}"

            ws = wb.create_sheet(title=sheet_name)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)
        print(f"  ✓ Progress saved: {len(tables)} tables")
    except Exception as e:
        print(f"  Warning: Could not save progress: {e}")


def create_excel_file(tables, output_path):
    """Create Excel file from extracted tables.

    Args:
        tables: List of table dictionaries with 'dataframe', 'page', and 'table' keys
        output_path: Path to save the Excel file

    Returns:
        Path: Path to the created Excel file
    """
    print(f"\nCreating Excel file with {len(tables)} table(s)...")

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    all_general_notes = []

    for idx, table_data in enumerate(tables, start=1):
        df = table_data['dataframe']
        page_num = table_data['page']
        table_num = table_data['table']

        # Identify rollup rows (for display purposes only, not for formulas)
        rollup_rows = identify_rollup_rows(df)
        if rollup_rows:
            print(f"  Identified {len(rollup_rows)} rollup row(s) with total values")

        # Extract general notes before creating sheet
        general_notes = extract_general_notes(df)
        if general_notes:
            all_general_notes.extend([(page_num, note) for note in general_notes])

        if len(tables) == 1:
            sheet_name = "Sheet1"
        else:
            sheet_name = f"Page{page_num}_Table{table_num}"

        if len(sheet_name) > 31:
            sheet_name = f"P{page_num}_T{table_num}"

        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name}_{idx}"

        ws = wb.create_sheet(title=sheet_name)

        # Write data to worksheet (with original values, no formulas)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        print(f"  Sheet {idx}: {sheet_name} (Page {page_num}, {len(df)} rows x {len(df.columns)} columns)")

    # Create separate Notes tab if there are general notes
    if all_general_notes:
        notes_ws = wb.create_sheet(title="Notes")
        notes_ws.cell(row=1, column=1, value="Page")
        notes_ws.cell(row=1, column=2, value="Note")
        for note_idx, (page_num, note) in enumerate(all_general_notes, start=2):
            notes_ws.cell(row=note_idx, column=1, value=f"Page {page_num}")
            notes_ws.cell(row=note_idx, column=2, value=note)
        print(f"  ✓ Created Notes tab with {len(all_general_notes)} general note(s)")

    wb.save(output_path)
    print(f"✓ Successfully created: {output_path}\n")

    return output_path
