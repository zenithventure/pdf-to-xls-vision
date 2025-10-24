#!/usr/bin/env python3
"""
PDF to XLS Converter using Claude Vision API
Handles both text-based and image-based (scanned) PDFs using AI.
"""

import os
import sys
import argparse
from pathlib import Path
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import anthropic
import base64
from io import BytesIO
import json
from dotenv import load_dotenv
import fitz  # PyMuPDF
from PIL import Image, ImageOps
import pytesseract

# Load environment variables from .env file
load_dotenv()


def get_api_key():
    """Get Anthropic API key from environment."""
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key or api_key == 'your-api-key-here':
        raise ValueError(
            "ANTHROPIC_API_KEY not found or not set.\n"
            "Please edit the .env file and add your API key.\n"
            "Get your API key from: https://console.anthropic.com/"
        )
    return api_key


def get_model_name():
    """Get Claude model name from environment."""
    model = os.environ.get('CLAUDE_MODEL', 'claude-sonnet-4-5-20250929')
    return model


def pdf_has_text(pdf_path):
    """Check if PDF has extractable text or is image-based."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:3]:  # Check first 3 pages
                text = page.extract_text()
                if text and len(text.strip()) > 50:
                    return True
        return False
    except:
        return False


def pdf_is_image_based(pdf_path):
    """Check if PDF is image-based (contains images but may also have OCR'd text)."""
    try:
        doc = fitz.open(pdf_path)
        for page_num in range(min(3, len(doc))):  # Check first 3 pages
            page = doc[page_num]
            # Check if page has images
            image_list = page.get_images()
            if image_list:
                # Has images - likely scanned/image-based
                doc.close()
                return True
        doc.close()
        return False
    except:
        return False


def _fix_cell_parens(value):
    """Fix common parenthesis issues in a single cell.

    Handles:
    - Spaces inside parens: "( 297)" -> "(297)"
    - Duplicate opening parens: "((123)" -> "(123)"
    - Missing closing paren: "( 4410" -> "(4410)"
    - Orphaned closing paren: "123)" -> "(123)"
    """
    import re

    if not isinstance(value, str):
        return value

    val = str(value).strip()

    # Pattern 1: Remove spaces after opening paren: "( 123)" -> "(123)"
    val = re.sub(r'\(\s+', '(', val)

    # Pattern 2: Remove spaces before closing paren: "(123 )" -> "(123)"
    val = re.sub(r'\s+\)', ')', val)

    # Pattern 3: Remove duplicate opening parens: "((123)" -> "(123)"
    val = re.sub(r'\(+', '(', val)

    # Pattern 4: Fix missing closing paren for negative numbers: "( 123" -> "(123)"
    # Only if it starts with ( and contains digits
    if val.startswith('(') and not val.endswith(')'):
        if re.search(r'[\d,.-]+$', val):
            val = val + ')'

    # Pattern 5: Fix orphaned closing paren: "123)" -> "(123)" if it looks like a negative number
    if val.endswith(')') and not val.startswith('('):
        # Check if it's a number followed by )
        if re.match(r'^[\d,.-]+\)$', val):
            val = '(' + val

    return val


def clean_malformed_parentheses(df):
    """Clean malformed parentheses within individual cells.

    This is a post-processing step that fixes common OCR errors from Claude Vision API:
    - Spaces inside parentheses
    - Double opening parentheses
    - Missing closing parentheses
    - Orphaned closing parentheses
    """
    for col in df.columns:
        df[col] = df[col].apply(lambda x: _fix_cell_parens(x) if pd.notna(x) else x)

    return df


def clean_dataframe_parentheses(df):
    """Clean misplaced parentheses in entire dataframe.

    Handles cascading typewriter artifacts where parentheses are shifted across multiple cells.
    A cell ending with '(' means that '(' belongs to the next cell. This can cascade across
    multiple cells in a row.

    Example cascade:
    Before: ["10,947 (", "3,094)(", "578)(", "173"]
    After:  ["10,947", "(3,094)", "(578)", "(173"]

    The logic:
    - Scan left to right looking for cells ending with '('
    - Move that '(' to the next cell
    - If next cell has ')(' pattern, split it: ) stays with current, number gets wrapped, ( cascades
    - Continue until no more trailing '(' found
    """
    import re

    # Process each row
    for idx in df.index:
        # Keep processing until no more changes are made
        changed = True
        while changed:
            changed = False

            for col_idx in range(len(df.columns) - 1):
                curr_col = df.columns[col_idx]
                next_col = df.columns[col_idx + 1]

                curr_val = df.at[idx, curr_col]
                next_val = df.at[idx, next_col]

                # Check if current cell ends with '('
                if pd.notna(curr_val):
                    curr_str = str(curr_val).strip()

                    if curr_str.endswith('('):
                        # Remove trailing '(' from current cell
                        curr_str = curr_str[:-1].strip()

                        # Process next cell
                        if pd.notna(next_val):
                            next_str = str(next_val).strip()

                            # Check if next cell has ')(' pattern
                            match = re.search(r'^([\d,.-]+)\)\($', next_str)
                            if match:
                                # Pattern: curr="X (" + next="123)("
                                # With incoming (, next becomes (123) with trailing ( to cascade
                                # Result: curr="X" + next="(123)("
                                number = match.group(1)
                                df.at[idx, curr_col] = curr_str if curr_str else None
                                df.at[idx, next_col] = f'({number})('
                                changed = True
                            elif next_str.endswith(')') and not next_str.startswith('('):
                                # Pattern: curr="X (" + next="123)"
                                # Result: curr="X" + next="(123)"
                                df.at[idx, curr_col] = curr_str if curr_str else None
                                df.at[idx, next_col] = f'({next_str}'
                                changed = True
                            else:
                                # Just move ( to beginning of next cell
                                df.at[idx, curr_col] = curr_str if curr_str else None
                                df.at[idx, next_col] = '(' + next_str
                                changed = True
                        else:
                            # Next cell is empty
                            df.at[idx, curr_col] = curr_str if curr_str else None
                            df.at[idx, next_col] = '('
                            changed = True

                # Also check if next cell has ')(' pattern without incoming '('
                # This handles cells like "3,094)(" where the ) should go to previous cell
                if pd.notna(next_val):
                    next_str = str(next_val).strip()
                    match = re.search(r'^([\d,.-]+)\)\($', next_str)
                    if match:
                        # Check if current cell doesn't end with '('
                        curr_str = str(curr_val).strip() if pd.notna(curr_val) else ''
                        if not curr_str.endswith('('):
                            # Pattern: curr="X" + next="123)("
                            # Move ) to curr, wrap number, keep trailing (
                            # Result: curr="X)" + next="(123)("
                            number = match.group(1)
                            df.at[idx, curr_col] = (curr_str + ')') if curr_str else ')'
                            df.at[idx, next_col] = f'({number})('
                            changed = True

    # Second pass: Clean up any remaining percentage artifacts
    # Pattern: "-3.34% (" should be "-3.34%"
    for col in df.columns:
        df[col] = df[col].apply(lambda x: re.sub(r'(%)\s*\($', r'\1', str(x).strip()) if pd.notna(x) and isinstance(x, str) else x)

    return df


def detect_orientation(img):
    """Detect image orientation using pytesseract OSD."""
    try:
        # Use pytesseract to detect orientation
        osd = pytesseract.image_to_osd(img)

        # Parse the OSD output to get rotation angle
        rotation = 0
        confidence = 0
        for line in osd.split('\n'):
            if 'Rotate:' in line:
                rotation = int(line.split(':')[1].strip())
            if 'Orientation confidence:' in line:
                confidence = float(line.split(':')[1].strip())

        return rotation, confidence
    except Exception as e:
        # If OSD fails, return 0 (no rotation)
        return 0, 0


def convert_pdf_page_to_image(pdf_path, page_num):
    """Convert a PDF page to base64-encoded image using PyMuPDF, with automatic orientation detection."""
    try:
        # Open PDF with PyMuPDF
        doc = fitz.open(pdf_path)

        # Get the page (0-indexed)
        page = doc[page_num - 1]

        # Render page to image at high resolution
        mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better quality
        pix = page.get_pixmap(matrix=mat)

        # Convert to PIL Image
        img_bytes = pix.tobytes("png")
        img = Image.open(BytesIO(img_bytes))

        # Close document
        doc.close()

        # Detect actual visual orientation using OCR
        detected_rotation, confidence = detect_orientation(img)

        # Apply rotation correction if needed (only if confidence is reasonable)
        if detected_rotation != 0 and confidence > 1.0:
            # Rotation direction conversion:
            # - Tesseract OSD "Rotate" value = degrees to rotate CLOCKWISE to correct orientation
            # - PIL's rotate() method rotates COUNTER-CLOCKWISE by default
            # - Therefore: PIL_angle = -Tesseract_angle to convert conventions
            # - expand=True ensures the canvas expands to fit the rotated image without cropping
            # Example: If text is 90° clockwise (sideways right), Tesseract returns 270,
            #          and rotate(-270) = rotate 90° clockwise = corrects the orientation
            img = img.rotate(-detected_rotation, expand=True)
            print(f"    Detected rotation {detected_rotation}° (confidence: {confidence:.1f}) - correcting")

        # Convert PIL Image to PNG bytes
        output = BytesIO()
        img.save(output, format='PNG')
        final_img_data = output.getvalue()

        # Encode to base64
        return base64.standard_b64encode(final_img_data).decode('utf-8')

    except Exception as e:
        print(f"    Error converting page to image: {e}")
        import traceback
        traceback.print_exc()
        return None


def extract_table_with_claude_vision(pdf_path, client, model_name, output_path=None, save_every=10):
    """Extract tables from PDF using Claude Vision API with incremental saving.

    Args:
        pdf_path: Path to PDF file
        client: Anthropic API client
        model_name: Claude model name
        output_path: Optional path to save incremental progress
        save_every: Save progress every N pages (default: 10)
    """
    print(f"  Using Claude Vision API ({model_name}) for extraction...")
    tables = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            num_pages = len(pdf.pages)

            for page_num in range(1, num_pages + 1):
                print(f"  Processing page {page_num}/{num_pages} with Claude Vision...")

                # Convert page to image
                image_data = convert_pdf_page_to_image(pdf_path, page_num)

                if not image_data:
                    print(f"    Could not convert page {page_num} to image")
                    continue

                # Call Claude Vision API
                try:
                    message = client.messages.create(
                        model=model_name,
                        max_tokens=4096,
                        messages=[
                            {
                                "role": "user",
                                "content": [
                                    {
                                        "type": "image",
                                        "source": {
                                            "type": "base64",
                                            "media_type": "image/png",
                                            "data": image_data,
                                        },
                                    },
                                    {
                                        "type": "text",
                                        "text": """Extract all tabular data from this image and return it as a CSV format.

Requirements:
1. Preserve all rows and columns exactly as they appear, including:
   - Total/summary rows with their FULL labels (e.g., "Total Other Income", "Gross Potential Income", "Effective Gross Income (EGI)", "Utilities Total", "Total Expenses", "Net Operating Income (NOI)")
   - ALL breakdown/sub-item rows (e.g., "Parking", "Utility Reimbursement", "Pet Fee")
   - Indented or hierarchical items must be included as separate rows
2. Keep all numbers, text, and formatting characters
3. Use commas to separate columns
4. Put values with commas inside quotes
5. Include column headers if present
6. Add a "Row_Type" column as the FIRST column to indicate the type of each row:
   - Use "ROLLUP" for rows that contain words like "Total", "Gross", "Net", "Effective" and represent sums (e.g., "Total Other Income", "Total Expenses", "Net Operating Income")
   - Use "DETAIL" for individual line items that are not totals
   - Use "HEADER" for header/section title rows
7. CRITICAL: Look for notes, annotations, or text outside/beside the main table columns:
   - If you see a "NOTES AND ASSUMPTIONS" section or numbered notes on the side, create a "Notes" column as the LAST column
   - Add the full text of each note to its corresponding row ONLY if the note specifically references that row
   - If a note is general context (not tied to a specific row), leave the Notes column empty for that row
   - Include ALL text content visible in the image, not just the numeric table data
8. Return ONLY the CSV data, no explanation

IMPORTANT:
- Do NOT skip breakdown items or sub-categories. Every line item visible in the table must appear in the output.
- Do NOT skip total/rollup rows. These are CRITICAL and must include their full labels with all numbers.
- Do NOT skip text annotations, notes, or explanatory text. All text content should be captured.
- Clearly mark which rows are ROLLUP totals vs DETAIL items using the Row_Type column.

If there are multiple tables, extract the largest/main table and any associated notes."""
                                    }
                                ],
                            }
                        ],
                    )

                    # Extract CSV from response
                    csv_content = message.content[0].text.strip()

                    # Remove markdown code blocks if present
                    if csv_content.startswith('```'):
                        lines = csv_content.split('\n')
                        csv_content = '\n'.join(lines[1:-1]) if len(lines) > 2 else csv_content

                    # Improved validation: Check if content is valid CSV with actual table data
                    if csv_content and csv_content.strip():
                        # Try to parse as CSV to validate it's actual table data
                        from io import StringIO
                        df = None
                        try:
                            # Try standard CSV parsing
                            df = pd.read_csv(StringIO(csv_content))
                        except Exception as e:
                            # If CSV parsing fails, try with on_bad_lines='skip'
                            try:
                                df = pd.read_csv(StringIO(csv_content), on_bad_lines='skip')
                            except:
                                # Last resort: try reading as TSV or with different settings
                                try:
                                    df = pd.read_csv(StringIO(csv_content), sep=None, engine='python')
                                except:
                                    print(f"    CSV parsing error on page {page_num}: {e}")
                                    continue

                        # Check if it has at least one cell of data
                        if df is not None and not df.empty and df.shape[0] > 0 and df.shape[1] > 0:
                            # Valid table - proceed with data cleaning

                            # Clean up
                            df = df.dropna(how='all').dropna(axis=1, how='all')

                            # Fix typewriter parenthesis artifacts (only after successful parsing)
                            # This fixes values like "3,094)(" -> "(3,094)"
                            try:
                                df = clean_dataframe_parentheses(df)
                            except Exception as e:
                                print(f"    Warning: Could not clean cascading parentheses: {e}")

                            # Fix malformed parentheses within individual cells
                            # This fixes OCR errors like "( 297)" -> "(297)" and "( 4410" -> "(4410)"
                            try:
                                df = clean_malformed_parentheses(df)
                            except Exception as e:
                                print(f"    Warning: Could not clean malformed parentheses: {e}")

                            if not df.empty and len(df) > 0:
                                tables.append({
                                    'dataframe': df,
                                    'page': page_num,
                                    'table': 1
                                })
                                print(f"    ✓ Extracted table: {len(df)} rows x {len(df.columns)} columns")
                            else:
                                print(f"    No valid table data on page {page_num}")
                        else:
                            print(f"    No table content found on page {page_num}")
                    else:
                        print(f"    No table content found on page {page_num}")

                except Exception as e:
                    print(f"    API error on page {page_num}: {e}")
                    continue

                # Save progress incrementally every N pages
                if output_path and save_every > 0 and len(tables) > 0 and len(tables) % save_every == 0:
                    _save_excel_incremental(tables, output_path, page_num, num_pages)

    except Exception as e:
        print(f"  Vision extraction failed: {e}")
        import traceback
        traceback.print_exc()

    return tables


def identify_rollup_rows(df):
    """Identify rollup rows based on naming patterns and Row_Type column."""
    rollup_indicators = ['total', 'gross', 'effective', 'net operating income', 'noi']
    rollup_rows = []

    # Check if Row_Type column exists
    if 'Row_Type' in df.columns:
        for idx in df.index:
            try:
                row_type = df.loc[idx, 'Row_Type']
                # Handle case where loc returns a Series
                if isinstance(row_type, pd.Series):
                    row_type = row_type.iloc[0] if len(row_type) > 0 else None
                if pd.notna(row_type) and str(row_type).strip().upper() == 'ROLLUP':
                    rollup_rows.append(idx)
            except (KeyError, IndexError):
                continue
    else:
        # Fallback: identify by naming patterns
        first_col = df.columns[0]
        for idx in df.index:
            try:
                value = df.loc[idx, first_col]
                # Handle case where loc returns a Series
                if isinstance(value, pd.Series):
                    value = value.iloc[0] if len(value) > 0 else None
                if pd.notna(value):
                    value_lower = str(value).lower().strip()
                    if any(indicator in value_lower for indicator in rollup_indicators):
                        rollup_rows.append(idx)
            except (KeyError, IndexError):
                continue

    return rollup_rows


def add_rollup_formulas(df, ws, rollup_rows):
    """Add Excel formulas for rollup rows that sum their component rows."""
    # Find label column names (usually first non-Row_Type column that contains text)
    label_cols = ['Row_Type', 'INCOME', 'Category', 'EXPENSES', 'Subcategory', 'Description', 'Item']

    # Find numeric columns (columns with currency/number patterns, exclude Row_Type, labels, and Notes)
    numeric_cols = []
    for col in df.columns:
        # Skip if it's a label column or Notes
        if col in label_cols or col in ['Notes', 'notes', 'Note']:
            continue
        # Check if column contains numeric-looking data (currency signs, numbers, percentages)
        sample_values = df[col].dropna().head(5).astype(str)
        if any(any(char in str(val) for char in ['$', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '%', ',']) for val in sample_values):
            numeric_cols.append(col)

    for rollup_idx in rollup_rows:
        # Find the range of detail rows
        # Strategy: Look forward first (for breakdowns that come AFTER the total row)
        # If none found, look backward (for totals that come AFTER detail rows)
        start_detail_idx = None
        end_detail_idx = None

        # Try looking forward first (for cases like "Total Other Income" followed by "Parking", "Utility", etc.)
        for i in range(rollup_idx + 1, len(df)):
            if 'Row_Type' in df.columns:
                row_type = df.at[i, 'Row_Type']
                if pd.notna(row_type):
                    row_type_str = str(row_type).strip().upper()
                    if row_type_str == 'DETAIL':
                        if start_detail_idx is None:
                            start_detail_idx = i
                        end_detail_idx = i
                    elif row_type_str in ['ROLLUP', 'HEADER']:
                        # Stop at next rollup or header
                        break

        # If no forward details found, look backwards (for cases like detail rows followed by "Total")
        if start_detail_idx is None:
            for i in range(rollup_idx - 1, -1, -1):
                if 'Row_Type' in df.columns:
                    row_type = df.at[i, 'Row_Type']
                    if pd.notna(row_type):
                        row_type_str = str(row_type).strip().upper()
                        if row_type_str == 'DETAIL':
                            if end_detail_idx is None:
                                end_detail_idx = i
                            start_detail_idx = i
                        elif row_type_str in ['ROLLUP', 'HEADER']:
                            # Stop at previous rollup or header
                            break

        if start_detail_idx is not None and end_detail_idx is not None:
            # Add formulas for each numeric column ONLY
            for col_idx, col_name in enumerate(df.columns):
                if col_name in numeric_cols:
                    # Excel row is +2 (1-indexed + header row)
                    excel_row = rollup_idx + 2
                    excel_start_row = start_detail_idx + 2
                    excel_end_row = end_detail_idx + 2

                    # Convert column index to Excel column letter
                    from openpyxl.utils import get_column_letter
                    excel_col = get_column_letter(col_idx + 1)

                    # Create SUM formula
                    formula = f"=SUM({excel_col}{excel_start_row}:{excel_col}{excel_end_row})"
                    ws.cell(row=excel_row, column=col_idx + 1, value=formula)


def extract_general_notes(df):
    """Extract notes that are not tied to specific rows (for separate Notes tab)."""
    general_notes = []

    if 'Notes' in df.columns:
        # Look for notes in header or standalone rows
        for idx in df.index:
            note = df.at[idx, 'Notes']
            if pd.notna(note):
                # Check if this is a general note (in a HEADER row or standalone)
                if 'Row_Type' in df.columns:
                    row_type = df.at[idx, 'Row_Type']
                    if pd.notna(row_type) and str(row_type).strip().upper() == 'HEADER':
                        general_notes.append(note)

    return general_notes


def _save_excel_incremental(tables, output_path, current_page, total_pages):
    """Save progress to Excel file incrementally."""
    try:
        print(f"  💾 Saving progress... ({current_page}/{total_pages} pages processed)")
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for idx, table_data in enumerate(tables, start=1):
            df = table_data['dataframe']
            page_num = table_data['page']
            table_num = table_data['table']

            if len(tables) == 1:
                sheet_name = "Sheet1"
            else:
                sheet_name = f"Page{page_num}_Table{table_num}"

            if len(sheet_name) > 31:
                sheet_name = f"P{page_num}_T{table_num}"

            if sheet_name in wb.sheetnames:
                sheet_name = f"{sheet_name}_{idx}"

            ws = wb.create_sheet(title=sheet_name)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)
        print(f"  ✓ Progress saved: {len(tables)} tables")
    except Exception as e:
        print(f"  Warning: Could not save progress: {e}")


def detect_quality_issues(table_data):
    """Detect signs of poor table extraction quality.

    Returns a list of detected quality issues (empty list if no issues).

    Quality heuristics:
    1. Single column trap: Flag if table has only 1 column with multiple rows
    2. Row explosion: Flag if row count is suspiciously high (>50 rows with other issues)
    3. Column consistency: Flag if >30% of rows have different column counts
    4. Empty cells ratio: Flag if >50% of cells are empty
    5. Duplicate rows: Flag if >20% of rows are duplicates
    6. Garbled text: Flag if >10% of cells contain encoding/parsing errors
    """
    issues = []

    if table_data is None or (isinstance(table_data, list) and len(table_data) == 0):
        return issues

    # Convert table_data to dataframe if it's a raw table
    if isinstance(table_data, list):
        try:
            df = pd.DataFrame(table_data[1:], columns=table_data[0])
        except:
            return ["Could not parse table data"]
    else:
        df = table_data

    if df.empty:
        return issues

    num_rows = len(df)
    num_cols = len(df.columns)

    # Heuristic 1: Single column trap with multiple rows (likely parsing failure)
    # Single column is only suspicious if there are multiple rows (>3)
    if num_cols == 1 and num_rows > 3:
        issues.append(f"Single column table with {num_rows} rows (likely parsing error)")

    # Heuristic 2: Row explosion (suspiciously high row count)
    # More nuanced: only flag if BOTH high row count AND other suspicious patterns
    row_explosion_detected = False
    if num_rows > 70:
        # Very high threshold - definitely suspicious
        issues.append(f"Excessive row count ({num_rows} rows, likely incorrect parsing)")
        row_explosion_detected = True
    elif num_rows > 50:
        # Medium threshold - only flag if combined with narrow columns or high emptiness
        # Check if columns are suspiciously narrow (suggests over-splitting)
        if num_cols > 12:
            issues.append(f"Excessive row count ({num_rows} rows) with many columns ({num_cols}), likely incorrect parsing")
            row_explosion_detected = True

    # Heuristic 3: Check column count consistency across rows
    # Count non-null values in each row as a proxy for effective column count
    row_col_counts = df.notna().sum(axis=1)
    if len(row_col_counts) > 0:
        most_common_count = row_col_counts.mode()[0] if len(row_col_counts.mode()) > 0 else num_cols
        inconsistent_rows = (row_col_counts != most_common_count).sum()
        inconsistency_ratio = inconsistent_rows / len(row_col_counts)

        if inconsistency_ratio > 0.3:
            issues.append(f"Inconsistent column counts ({inconsistency_ratio:.1%} of rows differ)")

    # Heuristic 4: Empty cells ratio
    # Only flag if table is large and mostly empty (suggests phantom rows/columns)
    total_cells = num_rows * num_cols
    empty_cells = df.isna().sum().sum()
    empty_ratio = empty_cells / total_cells if total_cells > 0 else 0

    # Stricter threshold for smaller tables, looser for larger ones
    empty_threshold = 0.6 if num_rows < 20 else 0.5
    if empty_ratio > empty_threshold:
        issues.append(f"High empty cell ratio ({empty_ratio:.1%} empty cells)")

    # Heuristic 5: Duplicate rows detection
    # pdfplumber often repeats rows when parsing fails
    if num_rows > 5:  # Only check tables with enough rows
        # Convert to string representation for comparison (handles NaN properly)
        df_str = df.astype(str)
        duplicated_rows = df_str.duplicated(keep='first').sum()
        duplicate_ratio = duplicated_rows / num_rows if num_rows > 0 else 0

        if duplicate_ratio > 0.2:  # More than 20% duplicates
            issues.append(f"High duplicate row ratio ({duplicate_ratio:.1%} of rows are duplicates, {duplicated_rows}/{num_rows} rows)")

    # Heuristic 6: Check for garbled text patterns (encoding issues, random characters)
    # Look for cells with unusual character patterns that suggest OCR/parsing errors
    import re
    garbled_count = 0
    sample_size = min(100, total_cells)  # Sample up to 100 cells
    cells_checked = 0

    for col in df.columns:
        for val in df[col].head(20):  # Check first 20 values per column
            if pd.notna(val) and isinstance(val, str):
                cells_checked += 1
                # Check for: excessive special chars, mixed scripts, control chars
                if re.search(r'[^\x20-\x7E\u00A0-\u024F\u20A0-\u20CF]{3,}', str(val)):  # Non-printable chars
                    garbled_count += 1
                elif len(val) > 5 and re.search(r'[^\w\s$,.%()\-\'/]{3,}', str(val)):  # Excessive special chars
                    garbled_count += 1

            if cells_checked >= sample_size:
                break
        if cells_checked >= sample_size:
            break

    if cells_checked > 0 and garbled_count / cells_checked > 0.1:
        issues.append(f"Garbled text detected ({garbled_count}/{cells_checked} cells)")

    return issues


def extract_tables_from_text_pdf(pdf_path):
    """Extract tables from text-based PDF using pdfplumber with quality validation.

    Returns:
        tuple: (tables, quality_issues_detected)
            - tables: List of extracted table dictionaries
            - quality_issues_detected: Boolean indicating if any quality issues were found
    """
    tables = []
    pages_with_issues = []
    all_quality_issues = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            # Try default extraction
            page_tables = page.extract_tables()

            # Try with text-based settings if nothing found
            if not page_tables:
                page_tables = page.extract_tables(table_settings={
                    "vertical_strategy": "text",
                    "horizontal_strategy": "text",
                    "snap_tolerance": 3,
                    "join_tolerance": 3,
                    "edge_min_length": 3,
                    "min_words_vertical": 3,
                    "min_words_horizontal": 1,
                })

            if page_tables:
                for table_num, table in enumerate(page_tables, start=1):
                    if table and len(table) > 0:
                        try:
                            df = pd.DataFrame(table[1:], columns=table[0])
                            df = df.dropna(how='all').dropna(axis=1, how='all')

                            if not df.empty:
                                # Check for quality issues
                                issues = detect_quality_issues(df)

                                if issues:
                                    pages_with_issues.append(page_num)
                                    all_quality_issues.extend(issues)
                                    print(f"  Page {page_num}, Table {table_num}: {len(df)} rows x {len(df.columns)} columns ⚠️  Quality issues detected")
                                    for issue in issues:
                                        print(f"    - {issue}")
                                else:
                                    print(f"  Page {page_num}, Table {table_num}: {len(df)} rows x {len(df.columns)} columns")

                                tables.append({
                                    'dataframe': df,
                                    'page': page_num,
                                    'table': table_num
                                })
                        except Exception as e:
                            print(f"  Warning: Could not process table on page {page_num}: {e}")

    quality_issues_detected = len(pages_with_issues) > 0

    if quality_issues_detected:
        print(f"\n  ⚠️  Quality issues detected on {len(set(pages_with_issues))} page(s)")

    return tables, quality_issues_detected


def convert_pdf_to_xls(pdf_path, output_path=None, output_dir=None, save_every=10, force_vision=False):
    """Convert PDF to Excel. Uses text extraction for text-based PDFs, Vision API with rotation detection for image-based PDFs.

    Args:
        pdf_path: Path to PDF file
        output_path: Optional output Excel file path
        output_dir: Optional output directory
        save_every: For large PDFs, save progress every N pages (default: 10)
        force_vision: Force Vision API extraction even for text-based PDFs (default: False)
    """
    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")

    if output_path:
        output_path = Path(output_path)
    else:
        if output_dir:
            output_dir = Path(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
        else:
            output_dir = pdf_path.parent
        output_path = output_dir / f"{pdf_path.stem}.xlsx"

    print(f"Converting: {pdf_path}")
    print(f"Output: {output_path}")

    try:
        # Check if PDF is image-based (scanned/photos)
        is_image_based = pdf_is_image_based(pdf_path)

        if force_vision or is_image_based:
            # Image-based PDF or forced Vision API: use Vision API with rotation detection
            if force_vision and not is_image_based:
                print("  Text-based PDF, using Vision API (forced)...")
            else:
                print("  Image-based PDF detected, using Vision API with rotation detection...")
            api_key = get_api_key()
            model_name = get_model_name()
            client = anthropic.Anthropic(api_key=api_key)
            tables = extract_table_with_claude_vision(pdf_path, client, model_name, output_path, save_every)
        else:
            # Text-based PDF: use direct extraction (fast, no API needed)
            print("  Text-based PDF, using direct extraction...")
            tables, quality_issues_detected = extract_tables_from_text_pdf(pdf_path)

            # Option 1: Auto-retry with Vision API if quality issues detected
            if quality_issues_detected:
                print("\n  ⚠️  Quality issues detected in text extraction!")
                print("  🔄 Retrying with Vision API for better accuracy...\n")
                api_key = get_api_key()
                model_name = get_model_name()
                client = anthropic.Anthropic(api_key=api_key)
                tables = extract_table_with_claude_vision(pdf_path, client, model_name, output_path, save_every)

        if not tables:
            print(f"Warning: No tables found in {pdf_path}")
            return None

        print(f"\nCreating Excel file with {len(tables)} table(s)...")

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        all_general_notes = []

        for idx, table_data in enumerate(tables, start=1):
            df = table_data['dataframe']
            page_num = table_data['page']
            table_num = table_data['table']

            # Identify rollup rows (for display purposes only, not for formulas)
            rollup_rows = identify_rollup_rows(df)
            if rollup_rows:
                print(f"  Identified {len(rollup_rows)} rollup row(s) with total values")

            # Extract general notes before creating sheet
            general_notes = extract_general_notes(df)
            if general_notes:
                all_general_notes.extend([(page_num, note) for note in general_notes])

            if len(tables) == 1:
                sheet_name = "Sheet1"
            else:
                sheet_name = f"Page{page_num}_Table{table_num}"

            if len(sheet_name) > 31:
                sheet_name = f"P{page_num}_T{table_num}"

            if sheet_name in wb.sheetnames:
                sheet_name = f"{sheet_name}_{idx}"

            ws = wb.create_sheet(title=sheet_name)

            # Write data to worksheet (with original values, no formulas)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            print(f"  Sheet {idx}: {sheet_name} (Page {page_num}, {len(df)} rows x {len(df.columns)} columns)")

        # Create separate Notes tab if there are general notes
        if all_general_notes:
            notes_ws = wb.create_sheet(title="Notes")
            notes_ws.cell(row=1, column=1, value="Page")
            notes_ws.cell(row=1, column=2, value="Note")
            for note_idx, (page_num, note) in enumerate(all_general_notes, start=2):
                notes_ws.cell(row=note_idx, column=1, value=f"Page {page_num}")
                notes_ws.cell(row=note_idx, column=2, value=note)
            print(f"  ✓ Created Notes tab with {len(all_general_notes)} general note(s)")

        wb.save(output_path)
        print(f"✓ Successfully created: {output_path}\n")

        return output_path

    except Exception as e:
        print(f"Error converting {pdf_path}: {e}")
        import traceback
        traceback.print_exc()
        raise


def batch_convert_directory(input_dir, output_dir=None, recursive=False, force_vision=False):
    """Batch convert PDFs in directory. Auto-detects text vs image-based PDFs.

    Args:
        input_dir: Directory containing PDF files
        output_dir: Optional output directory
        recursive: Recursively search subdirectories
        force_vision: Force Vision API extraction for all PDFs
    """
    input_dir = Path(input_dir)

    if not input_dir.exists():
        raise FileNotFoundError(f"Directory not found: {input_dir}")

    if recursive:
        pdf_files = list(input_dir.rglob("*.pdf"))
    else:
        pdf_files = list(input_dir.glob("*.pdf"))

    pdf_files = [f for f in pdf_files if ':Zone.Identifier' not in str(f)]

    if not pdf_files:
        print(f"No PDF files found in {input_dir}")
        return

    print(f"Found {len(pdf_files)} PDF file(s)")
    print("=" * 70)

    success_count = 0
    failed_files = []

    for pdf_path in pdf_files:
        try:
            if output_dir and recursive:
                rel_path = pdf_path.relative_to(input_dir)
                out_dir = Path(output_dir) / rel_path.parent
            else:
                out_dir = output_dir or pdf_path.parent

            result = convert_pdf_to_xls(pdf_path, output_dir=out_dir, force_vision=force_vision)
            if result:
                success_count += 1
            print("=" * 70)

        except Exception as e:
            print(f"Failed to convert {pdf_path}: {e}")
            failed_files.append(pdf_path)
            print("=" * 70)

    print(f"\n✓ Conversion complete!")
    print(f"  Successful: {success_count}/{len(pdf_files)}")

    if failed_files:
        print(f"\n✗ Failed files ({len(failed_files)}):")
        for f in failed_files:
            print(f"  - {f}")


def main():
    parser = argparse.ArgumentParser(
        description="PDF to Excel converter with auto-detection (text extraction or Vision API with rotation)"
    )
    parser.add_argument("input", help="PDF file or directory")
    parser.add_argument("-o", "--output", help="Output file or directory")
    parser.add_argument("-r", "--recursive", action="store_true",
                       help="Recursively search subdirectories")
    parser.add_argument("--force-vision", action="store_true",
                       help="Force Vision API extraction even for text-based PDFs (useful for complex table layouts)")

    args = parser.parse_args()
    input_path = Path(args.input)

    if not input_path.exists():
        print(f"Error: Path not found: {input_path}")
        sys.exit(1)

    # Check for API key
    try:
        get_api_key()
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)

    try:
        if input_path.is_file():
            convert_pdf_to_xls(input_path, output_path=args.output, force_vision=args.force_vision)
        elif input_path.is_dir():
            batch_convert_directory(input_path, output_dir=args.output,
                                   recursive=args.recursive, force_vision=args.force_vision)
        else:
            print(f"Error: Invalid path: {input_path}")
            sys.exit(1)

    except KeyboardInterrupt:
        print("\nConversion cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"\nError: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
